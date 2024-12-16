import pandas as pd
import openpyxl
import os
import sys

# カテゴリとテンプレートファイルのマッピング（優先順位付き）
CATEGORY_MAPPING = [
    ('HEL', 'ヘルメットグラフ作成.xlsm'),
    ('BICYCLE', '自転車帽グラフ作成.xlsm'),
    ('BASEBALL', '野球帽グラフ作成.xlsm'),
    ('FALLALL', '安全帯グラフ作成.xlsm')
]

def load_csv_files(directory):
    """指定ディレクトリから条件に合致するCSVファイルを読み込む"""
    csv_files = [os.path.join(directory, f) for f in os.listdir(directory)
                 if 'InspectionLOG' in f and f.lower().endswith('.csv')]
    if not csv_files:
        print("CSVファイルが見つかりません。")
    else:
        for file in csv_files:
            try:
                data = pd.read_csv(file)
                print(f"{file} 読み込み成功")
            except Exception as e:
                print(f"{file} 読み込み失敗: {e}")
        return pd.concat([pd.read_csv(file) for file in csv_files], ignore_index=True)
    return pd.DataFrame()

def filter_data(data_frame):
    """2列目に特定のキーワードが含まれるレコードをフィルタリングする"""
    keywords = [category for category, _ in CATEGORY_MAPPING]
    return data_frame[data_frame.iloc[:, 1].str.contains('|'.join(keywords), na=False)]

def get_sheet_name(value):
    """値に基づいて転記するシート名を決定する"""
    mapping = {
        'HEL': 'LOG_Helmet',
        'BICYCLE': 'LOG_Bicycle',
        'BASEBALL': 'LOG_BaseBall',
        'FALLALL': 'LOG_FallArrest'
    }
    for key, sheet_name in mapping.items():
        if key in value:
            return sheet_name
    return None

def get_prefix_from_sheet(sheet_name):
    """シート名からファイル名のプレフィックスを取得する"""
    prefix_mapping = {
        'LOG_Helmet': 'Helmet',
        'LOG_Bicycle': 'Bicycle',
        'LOG_BaseBall': 'BaseBall',
        'LOG_FallArrest': 'FallAll'
    }
    return prefix_mapping.get(sheet_name)

def get_template_file(filtered_data):
    """データの内容に基づいて使用するテンプレートファイルを決定する"""
    # カテゴリごとのデータ件数をカウント
    category_counts = {}
    for category, _ in CATEGORY_MAPPING:
        count = filtered_data[filtered_data.iloc[:, 1].str.contains(category, na=False)].shape[0]
        category_counts[category] = count

    # データが存在する場合
    if any(category_counts.values()):
        # 最大件数を取得
        max_count = max(category_counts.values())
        # 最大件数を持つカテゴリを優先順位に従って選択
        for category, template in CATEGORY_MAPPING:
            if category_counts[category] == max_count:
                return template, False  # (テンプレート名, 警告フラグ)

    # データが1件も無い場合
    return 'ヘルメットグラフ作成.xlsm', True

def validate_template(directory, template_file):
    """テンプレートファイルの存在を確認する"""
    template_path = os.path.join(directory, template_file)
    if not os.path.exists(template_path):
        print(f"読み込むテンプレートシート（{template_file}）が存在しません。")
        sys.exit(1)
    return template_path

def find_start_row(sheet, start_col):
    """転記を開始する行を見つける"""
    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        if sheet.cell(row=row, column=start_col).value is None:
            return row
    return max_row + 1

def write_data_to_sheet(sheet, data_frame, start_row, start_col):
    """データフレームの内容をExcelシートに転記する"""
    for index, row in enumerate(data_frame.iterrows(), start=0):
        for col_index, value in enumerate(row[1], start=0):
            sheet.cell(row=start_row + index, column=start_col + col_index - 1).value = value

def save_workbook(workbook, directory, sheet_names_used):
    """ワークブックを新しいファイル名で保存する"""
    prefixes = [get_prefix_from_sheet(sheet_name) for sheet_name in sheet_names_used]
    unique_prefixes = sorted(set(prefix for prefix in prefixes if prefix))

    if not unique_prefixes:
        base_filename = 'グラフ作成用ファイル'
    else:
        base_filename = '_'.join(unique_prefixes) + '_グラフ作成用ファイル'

    file_extension = '.xlsm'
    files = os.listdir(directory)
    count = 0
    pattern = f"{base_filename}_{{}}{file_extension}"

    while True:
        new_filename = pattern.format(count)
        if new_filename not in files:
            break
        count += 1

    new_path = os.path.join(directory, new_filename)
    workbook.save(new_path)
    workbook.close()
    print(f"データの転記が完了しました。保存先: {new_path}")
    return new_path

def main():
    current_directory = os.getcwd()
    parent_directory = os.path.dirname(current_directory)

    # データの処理
    df = load_csv_files(parent_directory)
    filtered_data = filter_data(df)
    if filtered_data.empty:
        print("条件に合致するデータがありません。")
        return

    # テンプレートファイルの選択と検証
    template_file, show_warning = get_template_file(filtered_data)
    template_path = validate_template(parent_directory, template_file)

    # Excelファイルを読み込む
    workbook = openpyxl.load_workbook(template_path, keep_vba=True)

    # 使用されたシート名を追跡
    sheet_names_used = set()

    # データを適切なシートに転記
    for index, row in filtered_data.iterrows():
        sheet_name = get_sheet_name(row.iloc[1])
        if sheet_name:
            sheet = workbook[sheet_name]
            sheet_names_used.add(sheet_name)
            start_row = find_start_row(sheet, 2)
            write_data_to_sheet(sheet, pd.DataFrame([row]), start_row, 2)
        else:
            print(f"警告: 対応するシートが見つかりませんでした。値: {row[1]}")

    # ワークブックを保存
    save_workbook(workbook, parent_directory, sheet_names_used)

    # 警告表示（データが1件も無い場合）
    if show_warning:
        print("警告: データが存在しないため、デフォルトテンプレート（ヘルメットグラフ作成.xlsm）を使用しました。")

if __name__ == '__main__':
    main()